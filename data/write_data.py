import xlrd
from openpyxl import load_workbook
from data.datapath import file_name


class WriteExcel():
    def __init__(self):
        self.data = xlrd.open_workbook(file_name, encoding_override=None)
        self.sheet = self.data.sheet_by_index(1)
        self.max_col = self.sheet.ncols

    @classmethod
    def write_excel_xls(cls, cell, value):
        """
        写入excel
        :param cell: 写入的单元格
        :param value: 写入的值
        :return:
        """
        workbook_ = load_workbook(file_name)
        sheetnames = workbook_.sheetnames  # 获得表单名字
        sheet = workbook_[sheetnames[1]]
        sheet[cell] = value
        workbook_.save(file_name)

    def get_connect_key(self, rows):
        """
        获取excel关键key值
        :param rows: 行号
        :return:
        """
        con_keys = self.sheet.cell_value(rows - 1, 1)
        return con_keys

    def get_json_value_by_key(self, in_json, target_key, results=[]):
        """
        根据key值读取对应的value值
        :param in_json:传入的json
        :param target_key: 目标key值
        :param results:
        :return:
        """
        if isinstance(in_json, dict):  # 如果输入数据的格式为dict
            for key in in_json.keys():  # 循环获取key
                data = in_json[key]
                self.get_json_value_by_key(data, target_key, results=results)  # 回归当前key对于的value
                if key == target_key:  # 如果当前key与目标key相同就将当前key的value添加到输出列表
                    results.append(data)
        elif isinstance(in_json, list) or isinstance(in_json, tuple):  # 如果输入数据格式为list或者tuple
            for data in in_json:  # 循环当前列表
                self.get_json_value_by_key(data, target_key, results=results)  # 回归列表的当前的元素
        return results

    def get_col(self):
        cells = []
        counts = []
        for i in range(1, self.max_col + 1):
            cell = self.sheet.cell_value(i, 0)
            cells.append(cell)
        for j in cells:
            count = cells.count(j)
            if count > 1:
                counts.append(j)
        return counts

    def get_key(self, rows):
        """
        获取接口返回值并写入excel
        :param cell: 写入的单元格地址，如C5
        :param rows: json存在的行号
        :return:
        """
        cell = self.sheet.cell_value(rows - 1, 0)
        key = self.get_connect_key(rows)
        json_data = self.sheet.cell_value(rows - 1, 3)
        if isinstance(json_data, str):
            da = self.get_json_value_by_key(eval(json_data), key)
            if len(da) > 0:
                if isinstance(da[0], list):
                    self.write_excel_xls(cell, da[0][0])
                else:
                    self.write_excel_xls(cell, da[0])
                return da
            else:
                print('未查到到{}值'.format(key))
                raise ValueError


if __name__ == '__main__':
    excel = WriteExcel()
    key = excel.get_key(3)
    print(key)
