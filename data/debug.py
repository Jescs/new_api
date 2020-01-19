
from openpyxl import load_workbook
from data.datapath import file_name


workbook_ = load_workbook(file_name)
print(workbook_.sheetnames)
sheetnames = workbook_.sheetnames  # 获得表单名字
sheet = workbook_[sheetnames[1]]
# sheet['A7'] = '47'
workbook_.save(file_name)
