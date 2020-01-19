from openpyxl import load_workbook
from data.datapath import file_name


def get_json_value_by_key(in_json, target_key, results=[]):
    """
    根据key值读取对应的value值
    :param in_json:
    :param target_key: 目标key值
    :param results:
    :return:
    """
    if isinstance(in_json, dict):
        for key in in_json.keys():  # 循环获取key
            data = in_json[key]
            get_json_value_by_key(data, target_key, results=results)  # 回归当前key对于的value
            if key == target_key:  # 如果当前key与目标key相同就将当前key的value添加到输出列表
                results.append(data)
    elif isinstance(in_json, list) or isinstance(in_json, tuple):  # 如果输入数据格式为list或者tuple
        for data in in_json:  # 循环当前列表
            get_json_value_by_key(data, target_key, results=results)  # 回归列表的当前的元素
    return results


data = """{
    "code": 10000,
    "msg": "ok",
    "data": [
        {
            "date": "2020-01-13",
            "period": [
                "pm"
            ]
        }
    ]
}"""

print(get_json_value_by_key(eval(data),'date'))


def write_excel_xls( cell, value):
    """
    写入excel
    :param cell: 写入的单元格
    :param value: 写入的值
    :return:
    """
    workbook_ = load_workbook(file_name)
    sheetnames = workbook_.sheetnames  # 获得表单名字
    sheet = workbook_[sheetnames[1]]
    sheet[cell] = value
    workbook_.save(file_name)

write_excel_xls('D2','111')