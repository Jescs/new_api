import xlrd
import os
from xlutils.copy import copy
from data.datapath import file_name
from datetime import datetime
from xlrd import xldate_as_tuple
from openpyxl import load_workbook

class ReadExcel:
    def __init__(self, num):
        """
        获取excel文件地址
        :param file_name:
        """
        self.num = num
        self.data_file = file_name
        self.data = xlrd.open_workbook(self.data_file)
        self.sheet = self.data.sheet_by_index(self.num)

    def write_excel_xls(self, col, row, value):
        """
        excel文件写入单元格
        :param num:
        :param col:
        :param row:
        :param value:
        :return:
        """
        book_w = copy(self.data)
        sheet_1 = book_w.get_sheet(self.num)
        sheet_1.write(row, col, value)
        os.remove(self.data_file)
        book_w.save(self.data_file)

    def get_max_row(self):
        """
        获取sheet最大行数
        :return:
        """
        rows = self.sheet.nrows
        return rows

    def get_max_col(self):
        """
        获取sheet最大列数
        :return:
        """
        return self.sheet.ncols

    def change_cell_type(self, row, col):
        ctype = self.sheet.cell(row, col).ctype  # 表格的数据类型
        cell = self.sheet.cell_value(row, col)
        if ctype == 2 and cell % 1 == 0:  # 如果是整形
            cell = int(cell)
        elif ctype == 3:
            # 转成datetime对象
            date = datetime(*xldate_as_tuple(cell, 0))
            cell = date.strftime('%Y/%m/%d')
        elif ctype == 4:
            cell = True if cell == 1 else False
        return cell

    def get_cell_value(self, row, col):
        """
        获取cell值
        :param col:
        :param row:
        :return:
        """
        return self.change_cell_type(row, col)

    def get_debug_values(self, row, col):
        return self.sheet.cell_value(row, col)

    def get_xlrd(self,row,col):
        wb = load_workbook(self.data_file, data_only=True)
        ws = wb.worksheets[self.num]
        return ws.cell(row,col).value


if __name__ == '__main__':
    excel = ReadExcel(0)
    print(excel.get_xlrd(4,13))
    # print(excel.get_debug_values(6, 4))
    # print(excel.change_cell_type(6, 4))
